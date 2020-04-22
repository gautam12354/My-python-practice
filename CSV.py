
#CSV Writing to a excel Book
# https://realpython.com/openpyxl-excel-spreadsheets-python/

'''
# Install package
from openpyxl import Workbook

# active sheet
workbook = Workbook()
sheet = workbook.active

# write data on sheet
sheet["A1"] = "hello"
sheet["B1"] = "world!"

# Save workbook
workbook.save(filename="hello_world.xlsx")

'''

# load package
from openpyxl import load_workbook

# load workbook
workbook = load_workbook(filename="sample.xlsx")
print(workbook.sheetnames)

# check active sheet
sheet = workbook.active
print(sheet)

# sheet Title
print(sheet.title)

# get A1 cell value
>>>sheet["A1"].value

>>>sheet["F10"].value

